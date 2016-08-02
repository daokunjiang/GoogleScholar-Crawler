# coding = utf-8

from selenium import webdriver
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook as ldwb
import re

def name_too_long(arg):

    # input may be None
    if not arg:
        return False

    try:
        t = re.findall('\.\.\.',arg)
    except TypeError as e:
        print(e)
        print(arg)
        print(str(arg))

    if len(t) == 0:
        return False

    print(arg)
    return True

def is_author_column(tag):
    return (tag.string == '作者' or tag.string == '发明者') and tag.has_attr('class') and (tag.get('class') == ['gsc_field'])

def is_magazine_column(tag):
    return (tag.string == '期刊' or tag.string == '研讨会论文') and tag.has_attr('class') and (tag.get('class') == ['gsc_field'])

def get_author_child(source):
    doc = bs(source, "lxml")
    x = doc.find(is_author_column)
    if not x:
        return ''
    x = x.find_next_sibling()
    name_return = x.string
    print(name_return)
    return name_return

def get_magazine_child(source):
    doc = bs(source, "lxml")
    x = doc.find(is_magazine_column)
    if not x:
        return ''
    x = x.find_next_sibling()
    print(x.string)
    return x.string


with open('sample_input.csv', 'r', encoding='gbk') as input_file:

    # read in every entry and do the work
    for x in input_file:
        x = x.split(',')
        name = x[0]
        url = x[1].strip()
        print(name)
        print(url)

        driver = webdriver.PhantomJS()
        driver.get(url)

        paper_name = []
        paper_author = []
        paper_magazine = []
        paper_ref = []
        paper_pubyear = []
        ref_year = []
        ref_num = []

        page_number = 1

        while 1:
            results = driver.page_source
            doc = bs(results, 'lxml')

            # get the html of the button for jumping to next page
            next_page_button = doc.find(id='gsc_bpf_next')
            print("Page " + str(page_number) + " : " + str(next_page_button.attrs))

            x = doc.find_all("tr", class_="gsc_a_tr")
            len_x = len(x)

            for i in range(0, len_x):

                # grab the title of the paper
                paper_name_raw = x[i].find(class_="gsc_a_at")
                paper_name_add = paper_name_raw.string
                paper_name.append(paper_name_add)

                # grab the authors of the paper
                paper_info_raw = x[i].find_all(class_="gs_gray")
                paper_author_add = paper_info_raw[0].string
                if not name_too_long(paper_author_add):
                    paper_author.append(paper_author_add)
                else:
                    url_child = paper_name_raw.get('href')
                    url_child = 'https://scholar.google.com' + url_child
                    dr = webdriver.PhantomJS()
                    dr.get(url_child)
                    source_child = dr.page_source
                    dr.quit()
                    paper_author.append(get_author_child(source_child))

                # grab the magazine of the paper
                magazine_raw = paper_info_raw[1]
                magazine_content = magazine_raw.contents
                if len(magazine_content) == 0:
                    paper_magazine.append('')
                else:
                    paper_magazine_add = magazine_content[0].string
                    if not name_too_long(paper_magazine_add):
                        paper_magazine.append(paper_magazine_add)
                    else:
                        url_child = paper_name_raw.get('href')
                        url_child = 'https://scholar.google.com' + url_child
                        dr = webdriver.PhantomJS()
                        dr.get(url_child)
                        source_child = dr.page_source
                        dr.quit()
                        paper_magazine.append(get_magazine_child(source_child))

                # grab the reference number of the paper
                paper_ref_raw = x[i].find_all(class_="gsc_a_c")
                paper_ref_a = paper_ref_raw[0].contents[0].string
                if paper_ref_a == '\xa0':
                    paper_ref.append('0')
                else:
                    paper_ref.append(paper_ref_a)

                # grab the pusblish year of the paper
                paper_pubyear_raw = x[i].find_all(class_="gsc_a_y")
                if paper_pubyear_raw[0].string == None:
                    paper_pubyear.append('')
                else:
                    paper_pubyear.append(paper_pubyear_raw[0].string)

            # if this button is disabled, then read the bar chart and break
            if next_page_button.attrs.get('disabled') == '':
                ref_year_raw = doc.find_all(id="gsc_g_x")[0]
                for year in ref_year_raw.children:
                    ref_year.append(year.string)

                ref_num_raw = doc.find_all(id="gsc_g_bars")[0]
                for num in ref_num_raw.children:
                    ref_num.append(num.string)

                break

            driver.find_element_by_id('gsc_bpf_next').click()
            page_number += 1

        print(paper_name)
        print(paper_author)
        print(paper_magazine)
        print(paper_ref)
        print(paper_pubyear)
        print(ref_year)
        print(ref_num)

        wb = ldwb(filename='sample_output3.xlsx')

        ws = wb.create_sheet(title=name)
        ws.append(['文章标题','作者','发表出处','引用数','发表年份'])
        for i in range(1,len(paper_name)+1):
            ws.cell(row=i + 1, column=1).value = paper_name[i-1]
            ws.cell(row=i + 1, column=2).value = paper_author[i-1]
            ws.cell(row=i + 1, column=3).value = paper_magazine[i-1]
            ws.cell(row=i + 1, column=4).value = paper_ref[i-1]
            ws.cell(row=i + 1, column=5).value = paper_pubyear[i-1]
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 80

        ref_name = name + '引用'
        ws_ref = wb.create_sheet(title=ref_name)
        ws_ref.append(ref_year)
        ws_ref.append(ref_num)

        wb.save(filename='sample_output3.xlsx')