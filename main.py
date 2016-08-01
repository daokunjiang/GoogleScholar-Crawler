# coding = utf-8

from selenium import webdriver
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook

with open('sample_input.csv', 'r', encoding='gbk') as input_file:
    wb = Workbook()
    for x in input_file:
        x = x.split(',')
        name = x[0]
        url = x[1].strip()
        print(name)
        print(url)

        driver = webdriver.PhantomJS()
        driver.get(url)

        paper_name = []
        paper_author = []
        paper_magazine = []
        paper_ref = []
        paper_pubyear = []
        ref_year = []
        ref_num = []

        while 1:
            results = driver.page_source
            doc = bs(results, 'lxml')

            # get the html of the button for jumping to next page
            next_page_button = doc.find(id='gsc_bpf_next')
            print("next_page_button.attrs: " + str(next_page_button.attrs))

            x = doc.find_all("tr", class_="gsc_a_tr")
            len_x = len(x)

            for i in range(0, len_x):
                paper_name_raw = x[i].find_all(class_="gsc_a_at")
                paper_name.append(paper_name_raw[0].string)

                paper_info_raw = x[i].find_all(class_="gs_gray")
                paper_author.append(paper_info_raw[0].string)
                magazine_raw = paper_info_raw[1]
                magazine_content = magazine_raw.contents
                if len(magazine_content) == 0:
                    paper_magazine.append('')
                else:
                    paper_magazine.append(magazine_content[0].string)

                paper_ref_raw = x[i].find_all(class_="gsc_a_c")
                paper_ref_a = paper_ref_raw[0].contents[0].string
                if paper_ref_a == '\xa0':
                    paper_ref.append('0')
                else:
                    paper_ref.append(paper_ref_a)
                paper_pubyear_raw = x[i].find_all(class_="gsc_a_y")
                if paper_pubyear_raw[0].string == None:
                    paper_pubyear.append('')
                else:
                    paper_pubyear.append(paper_pubyear_raw[0].string)

            # if this button is disabled, then read the bar chart and break
            if next_page_button.attrs.get('disabled') == '':
                ref_year_raw = doc.find_all(id="gsc_g_x")[0]
                for year in ref_year_raw.children:
                    ref_year.append(year.string)

                ref_num_raw = doc.find_all(id="gsc_g_bars")[0]
                for num in ref_num_raw.children:
                    ref_num.append(num.string)

                break

            driver.find_element_by_id('gsc_bpf_next').click()

        print(paper_name)
        print(paper_author)
        print(paper_magazine)
        print(paper_ref)
        print(paper_pubyear)
        print(ref_year)
        print(ref_num)

        ws = wb.create_sheet(title=name)
        ws.append(['文章标题','作者','发表出处','引用数','发表年份'])
        for i in range(1,len(paper_name)+1):
            ws.cell(row=i + 1, column=1).value = paper_name[i-1]
            ws.cell(row=i + 1, column=2).value = paper_author[i-1]
            ws.cell(row=i + 1, column=3).value = paper_magazine[i-1]
            ws.cell(row=i + 1, column=4).value = paper_ref[i-1]
            ws.cell(row=i + 1, column=5).value = paper_pubyear[i-1]
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 80

        ref_name = name + '引用'
        ws_ref = wb.create_sheet(title=ref_name)
        ws_ref.append(ref_year)
        ws_ref.append(ref_num)

    wb.save(filename='sample_output.xlsx')

# 现在能从成功的从主页抓数据了
# 下一阶段考虑深入到每一个文章页面去

# f = open('xue.txt', 'w', encoding='utf-8')
# re_len_before = 0
# count = 0
# # while 1:
# #     # results = driver.page_source
# #     # re_len = len(results)
# #     # print(re_len)
# #     # if (re_len_before != 0) & (re_len == re_len_before):
# #     #     break
# #     # re_len_before = re_len
# #     count += 1
# #     try:
# #         inputBtn = driver.find_element_by_id('gsc_bpf_more').click()
# #         driver.implicitly_wait(3)
# #     except EnvironmentError:
# #         break
#
# print(count)
#
# driver.implicitly_wait(10)
# results = driver.page_source
# f.write(results)

# inputBtn.send_keys('bing xie')
# driver.find_element_by_id('gs_hp_tsb').click()
# driver.find_element_by_xpath('//*[@id="gs_ccl"]/div[2]/table/tbody/tr/td[2]/h4/a').click()